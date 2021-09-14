from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtSql import QSqlRelationalDelegate

from customers.ui.CustomWidget.customAddOfferFrame import CustomAddOfferFrame
from customers.ui.CustomWidget.customAddOrderFrame import CustomAddOrderFrame
from customers.ui.OrderDialog.orderDialogMain import OrderDialog
from customers.ui.LoginDialog.LoginMain import LoginDialog
from customers.threads import ReportsWorker, Worker
from customers.ui.MainUI.ui_mainUI import Ui_MainWindow
from customers.ui.ProgressBarDialog.ProgressBarMain import PorgressBarDialog
from customers.ui.PriceDialog.PriceDialogMain import PriceDialog


import datetime
import os
import re
import sip
import openpyxl



from customers.database import changeSubsCost, checkFeesValue, checkShiftActive, copyData, finishShift, resetCounting, retrieveDailyNames, retrieveDailySubsState, retrieveItemAvailabelQuantity, retrieveItemId,retrieveItemNames, retrieveItemPrice, retrieveItemType, retrieveMonthlyNames, retrieveMonthlySubsState, retrieveMonthlySubsType, retrieveMonthlyid, retrieveOfferPrice, retrieveOrderType, retrieveEmployeesJobType, retrieveEmployeesNames, startShift, updateCurrentItemsQuantities, updateReports
from customers.model import *

# from PyQt5.uic import loadUiType

#  get the directory of this script
# path = os.path.dirname(os.path.abspath(__file__))

# MainWindowUI,MainWindowBase  = loadUiType(
#     os.path.join(path, 'Study_main.ui'))

# MainWindowUI = Ui_MainWindow
# MainWindowBase = QtWidgets.QMainWindow



class CustomersMainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self, daily_conn, archive_conn, supervisor_job_type, parent=None):
        QtWidgets.QMainWindow.__init__(self,parent)

        self.setupUi(self)

        self.supervisor_job_type = supervisor_job_type
        self.offer_id = None
        
        # DB connections
        self.daily_conn = daily_conn
        self.archive_conn = archive_conn
        
        
        # Flags
        self._DAILY_TABLES_FLAG = False
        self._ARCHIVE_TABLES_FLAG = False
        self._SETTINGS_TABLES_FLAG = False

        # Create model obejects 
        self.daily_customers_model = None
        self.monthly_customers_model = None
        self.orders_model = None
        self.warehouse_model = None
        self.employees_model = None
        self.reports_model = None
        self.offers_model = None
        self.shifts_model = None

        # Create sort model objects
        self.daily_customers_sort_model = None
        self.monthly_customers_sort_model = None
        self.order_sort_model = None
        self.warehouse_sort_model = None
        self.employees_sort_model = None
        self.offers_sort_model = None
        self.shifts_sort_model = None
         

        # Initial var for order numbering 
        self._order_item_number = 0
        self._employee_number = 0
        self._item_number = 0

        # insert an empty order and offer and employee previously
        self.plusOrder()
        self.plusOffer()

        # Apply desire changing to Main widnow
        self.uiCahnges()
        self.handleSignals()
        self.Completers()
        self.regexValidation()
        
        
    # def enterEvent(self, e):
    #     print(e.type())
    #     self.toggleMenuMaxWidth(250,True)

    
    # def keybords(self):
        # modifier = QtWidgets.QApplication.queryKeyboardModifiers()
        # if(modifier == Qt.ShiftModifier):
        #     self.toggleMenuMaxWidth(250,True)

    def closeEvent(self, event):
        """UI close envet handler"""

        title = None
        text = None

        # Check if any shift is started or not
        if(self.isAnyShiftActive(exit_event = True)):
            title = "Shift active alert"
            text = "There is still a shift is running\nConfirm exiting"
        
        else:
            title = "Closing window"
            text = "Confirm"

        
        reply = QtWidgets.QMessageBox.question(
            self,
            title,
            text,
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
        )
        if reply == QtWidgets.QMessageBox.Yes:
            LoginDialog(self.daily_conn)
            event.accept()
        else:
            event.ignore()
 
    def uiCahnges(self):
        """UI changes after run the program"""
        self.showMaximized()
        
        # UI changes in login
        # self.setWindowFlag(Qt.FramelessWindowHint)

        self.tabWidget.tabBar().setVisible(False)
        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'Main_tab'))

        # Resize tables view header sections into contentes
        self.daily_customers_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.daily_customers_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.daily_customers_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)

        
        self.monthly_customers_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.monthly_customers_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.monthly_customers_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)

        self.orders_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.orders_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.orders_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)
        
        self.warehouse_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.warehouse_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.warehouse_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)
        
        self.employees_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.employees_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.employees_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)
        
        self.reports_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeToContents)
        self.reports_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.reports_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.reports_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)
        # self.reports_tableView.resizeColumnsToContents()


        self.shifts_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.shifts_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.shifts_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)
           
        self.offers_tableView.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.offers_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignHCenter)
        self.offers_tableView.verticalHeader().setDefaultAlignment(Qt.AlignHCenter)


        # Add shadow to Logo image
        shadow = QtWidgets.QGraphicsDropShadowEffect(self, blurRadius=200,color=QtGui.QColor(255, 255, 255, 150 * 1), xOffset=3, yOffset=3)
        self.logo_frame.setGraphicsEffect(shadow)

        # Apply Blur on Main_tab frame
        self.blur = QtWidgets.QGraphicsBlurEffect(self)
        self.blur.setBlurRadius(1)  # Set blur radius to 1 to prevent the error that comes from Qpainter
        self.Main_tab.setGraphicsEffect(self.blur)



        # Change toolTip
        QtWidgets.QToolTip.setFont(QtGui.QFont("Times", 10, QtGui.QFont.Bold))


        # Adapt enabling buttons to supervisor type
        self.daily_customer_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        # self.daily_customer_export_btn.setEnabled(self.supervisor_job_type == 'Manager')
        # self.daily_customer_edit_price_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.order_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        # self.order_export_btn.setEnabled(self.supervisor_job_type == 'Manager')
        # self.monthly_customer_edit_cost_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.monthly_customer_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        # self.monthly_customer_edit_cost_btn.setEnabled(self.supervisor_job_type == 'Manager')
        # self.monthly_customer_export_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.warehouse_item_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.report_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.shift_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.employee_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
        self.offer_remove_btn.setEnabled(self.supervisor_job_type == 'Manager')
       
    def regexValidation(self):
        """Apply regular expression to some UI elements"""
        # Arabic names regular expression
        validator = QtGui.QRegularExpressionValidator(QtCore.QRegularExpression("[\s|ء-ي]+"))
        self.daily_customer_name_txt.setValidator(validator)
        self.daily_customer_name_filter_txt.setValidator(validator)
        self.monthly_customer_name_txt.setValidator(validator)
        self.monthly_customer_name_filter_txt.setValidator(validator)
        self.customer_name_txt2.setValidator(validator)
        self.orders_customer_name_filter_txt.setValidator(validator)
        self.employee_name_txt.setValidator(validator)

        # English names regular expression
        validator = QtGui.QRegularExpressionValidator(QtCore.QRegularExpression("[a-zA-Z]+"))
        self.employee_username_txt.setValidator(validator)


        # password regular expression
        validator = QtGui.QRegularExpressionValidator(QtCore.QRegularExpression("[\w+|\.]+"))
        self.employee_password_txt.setValidator(validator)

        # numbers regular expression
        validator = QtGui.QRegularExpressionValidator(QtCore.QRegularExpression("[0-9]+"))
        self.daily_customer_monthID_txt.setValidator(validator)
        self.warehouse_item_price_txt.setValidator(validator)
        self.warehouse_item_quantity_txt.setValidator(validator)
        # self.order_quantity_txt.setValidator(validator)
    
    def initialComboBoxs(self, db):
        self.daily_customer_subsType_filter_comboBox.clear()
        self.daily_customer_subsType_filter_comboBox.addItems(['']+ retrieveDailySubsState(db))
        self.daily_customer_subsType_filter_comboBox.setCurrentIndex(0)

        self.monthly_customer_subsType_filter_comboBox.clear()
        self.monthly_customer_subsType_filter_comboBox.addItems(['']+ retrieveMonthlySubsType(db))
        self.monthly_customer_subsType_filter_comboBox.setCurrentIndex(0)

        self.monthly_customer_subsState_filter_comboBox.clear()
        self.monthly_customer_subsState_filter_comboBox.addItems(['']+ retrieveMonthlySubsState(db))
        self.monthly_customer_subsState_filter_comboBox.setCurrentIndex(0)
        
        self.orders_type_filter_comboBox.clear()
        self.orders_type_filter_comboBox.addItems(['']+ retrieveOrderType(db))
        self.orders_type_filter_comboBox.setCurrentIndex(0)

    def initialSettingsComboBoxs(self):
        self.warehouse_item_type_filter_comboBox.clear()
        self.warehouse_item_type_filter_comboBox.addItems(['']+ retrieveItemType(self.daily_conn))

        self.employees_job_type_filter_comberoBox.clear()
        self.employees_job_type_filter_comberoBox.addItems(['']+ retrieveEmployeesJobType(self.daily_conn))

    def setCurrentDate(self):
        self.shifts_date_filter_dateEdit1.setDate(QtCore.QDate.currentDate())
        self.offers_date_filter_dateEdit1.setDate(QtCore.QDate.currentDate())

        self.shifts_sort_model.setDateFilter(self.shifts_date_filter_dateEdit1.date())
        self.shifts_sort_model.setDateFilter(self.offers_date_filter_dateEdit1.date())

    def Completers(self):
        """Add completer to some UI elements"""
        def addCompleter(data):

            # Get monthly names and add them to daily_customer_name_txt
            completer = QtWidgets.QCompleter(data,self)
            completer.popup().setItemDelegate(QtWidgets.QStyledItemDelegate(self))
            completer.popup().setObjectName("completerPopup")
            # completer.popup().setAutoFillBackground(True)
            completer.popup().setStyleSheet("""
            QAbstractItemView#completerPopup
            {
            font-size:20px;
            border: 4px solid rgb(150,150,150);
            padding: 2px;
            border-radius: 2px;
            outline: 0;
            }

            QAbstractItemView#completerPopup::item
            {
            
            border: 1px solid rgb(150,150,150);
            padding-top: 5px;
            padding-bottom: 8px;
            border-radius: 8px;
            }
            QAbstractItemView#completerPopup::item:hover
            {
            background-color:lightGrey;
            
            }
            
            """)
            completer.setCompletionMode(QtWidgets.QCompleter.PopupCompletion)
            completer.setCaseSensitivity(False)
            completer.setFilterMode(Qt.MatchContains)
            return completer
            
        self.daily_customer_name_txt.setCompleter(addCompleter(retrieveMonthlyNames(db  = self.daily_conn)))
        self.customer_name_txt2.setCompleter(addCompleter(retrieveDailyNames(db =self.daily_conn)))

    def handleSignals(self):
        """ handle all buttons """

        #Main window button
        self.archive_btn.clicked.connect(self.archivePanel)
        self.daily_btn.clicked.connect(self.dailyPanel)
        self.settings_btn.clicked.connect(self.settingsPanel)

        self.logout_btn.clicked.connect(self.BackToMain)
        self.exit_btn.clicked.connect(self.close)

        self.menu_btn.clicked.connect(lambda : self.toggleMainButtonsMenu(self.main_buttons_frame,300,True))

        # tabWidget buttons
        self.daily_customers_btn.clicked.connect(self.showDailyCustomers)
        self.monthly_subscrib_btn.clicked.connect(self.showMonthlyCustomers)
        self.orders_btn.clicked.connect(self.showOrders)
        self.reports_btn.clicked.connect(self.showReports)
        self.employees_btn.clicked.connect(self.showEmployees)
        self.warehouse_btn.clicked.connect(self.showWarehouse)
        self.offers_btn.clicked.connect(self.showOffers)
        self.shifts_btn.clicked.connect(self.showShifts)



        # Daily customers tab buttons
        self.daily_customer_add_btn2.clicked.connect(lambda: self.toggleMenuMaxWidth(self.frame_19,500,True))
        self.daily_customer_add_btn.clicked.connect(self.addDailyCustomer)
        self.daily_customer_remove_btn.clicked.connect(self.removeDailyCustomer)
        self.daily_customer_export_btn.clicked.connect(lambda : self.exportTable(self.daily_customers_model))
        self.daily_customer_add_order.clicked.connect(self.addOrderDirectly)
        self.daily_customer_search_btn.clicked.connect(lambda : self.toggleMenuMaxWidth(self.date_treeview_panel, 300, True))
        self.daily_customer_monthID_txt.textChanged['QString'].connect(self.setDailyName)
        self.daily_customer_name_txt.textChanged['QString'].connect(self.setDailyId)
        self.daily_customer_name_filter_txt.textChanged['QString'].connect(lambda customer_name: self.daily_customers_sort_model.setCustomerNameFilter(customer_name.strip()))
        self.daily_customer_subsType_filter_comboBox.currentTextChanged['QString'].connect(lambda sub_state: self.daily_customers_sort_model.setSubsStateFilter(sub_state))
        self.daily_customer_clear_btn.clicked.connect(self.clearDailyFilters)

        
        # Monthly customers tab buttons
        self.monthly_customer_add_btn2.clicked.connect(lambda: self.toggleMenuMaxWidth(self.frame_21,500,True))
        self.monthly_customer_add_btn.clicked.connect(self.addMonthlyCustomer)
        self.monthly_customer_remove_btn.clicked.connect(self.removeMonthlyCustomer)
        self.monthly_customer_update_btn.clicked.connect(self.updateMonthlyCustomer)
        self.monthly_customer_export_btn.clicked.connect(lambda : self.exportTable(self.monthly_customers_model))
        self.monthly_customer_search_btn.clicked.connect(lambda : self.toggleMenuMaxWidth(self.date_treeview_panel, 300, True))
        self.monthly_customer_name_filter_txt.textChanged['QString'].connect(lambda customer_name : self.monthly_customers_sort_model.setCustomerNameFilter(customer_name.strip()))
        self.monthly_customer_subsState_filter_comboBox.currentTextChanged['QString'].connect(lambda subs_state : self.monthly_customers_sort_model.setSubsStateFilter(subs_state))
        self.monthly_customer_subsType_filter_comboBox.currentTextChanged['QString'].connect(lambda subs_type : self.monthly_customers_sort_model.setSubsTypeFilter(subs_type))
        self.monthly_customer_clear_btn.clicked.connect(self.clearMonthlyFilters)

        # Orders tab buttons
        self.order_add_btn2.clicked.connect(lambda: self.toggleMenuMaxWidth(self.frame_11,600,True))
        self.order_add_btn.clicked.connect(self.addOrder)
        self.order_remove_btn.clicked.connect(self.removeOrder)
        self.order_export_btn.clicked.connect(lambda : self.exportTable(self.orders_model))
        self.plus_order_btn.clicked.connect(self.plusOrder)
        self.order_search_btn.clicked.connect(lambda : self.toggleMenuMaxWidth(self.date_treeview_panel, 300, True))
        self.orders_customer_name_filter_txt.textChanged['QString'].connect(lambda customer_name : self.orders_sort_model.setCustomerNameFilter(customer_name.strip()))
        self.orders_type_filter_comboBox.currentTextChanged['QString'].connect(lambda item_type : self.orders_sort_model.setItemTypeFilter(item_type))
        self.order_clear_btn.clicked.connect(self.clearOrderFitlers)
        self.order_sell_type_comboBox.currentTextChanged['QString'].connect(self.setTotalOrderPrice)

        # Warehouse tab buttons
        self.warehouse_item_add_btn2.clicked.connect(lambda : self.toggleMenuMaxWidth(self.frame_24,500, True))
        self.warehouse_item_add_btn.clicked.connect(self.addItem)
        self.warehouse_item_remove_btn.clicked.connect(self.removeItem)
        self.warehouse_update_current_items_btn.clicked.connect(self.updateItemsQauntities)
        self.warehouse_item_name_filter_txt.textChanged['QString'].connect(lambda item_name : self.warehouse_sort_model.setItemNameFilter(item_name.strip()))
        self.warehouse_item_type_filter_comboBox.currentTextChanged['QString'].connect(lambda item_type : self.warehouse_sort_model.setItemTypeFilter(item_type))
        self.warehouse_clear_btn.clicked.connect(self.clearWarehouseFitlers)


        # Subscription prices tab buttons
        self.daily_customer_edit_price_btn.clicked.connect(self.openPriceDialog)
        self.monthly_customer_edit_cost_btn.clicked.connect(self.openPriceDialog)
        self.order_show_btn.clicked.connect(self.openOrderDialog)

        # Supervisors tab buttons
        self.employee_add_btn2.clicked.connect(lambda : self.toggleMenuMaxWidth(self.frame_22,500,True))
        self.employee_add_btn.clicked.connect(self.addEmployees)
        self.employee_remove_btn.clicked.connect(self.removeEmployees)
        self.employees_employee_name_filter_txt.textChanged['QString'].connect(lambda employee_name : self.employees_sort_model.setEmployeeNameFilter(employee_name.strip()))
        self.employees_job_type_filter_comberoBox.currentTextChanged['QString'].connect(lambda job_type : self.employees_sort_model.setJobTypeFilter(job_type))
        self.employees_clear_btn.clicked.connect(self.clearEmployeesFilters)

        # Shifts tab butons
        self.shift_add_btn2.clicked.connect(lambda: self.toggleMenuMaxWidth(self.frame_69,500,True))
        self.shift_add_btn.clicked.connect(self.addShift)
        self.plus_employee_btn.clicked.connect(self.plusEmployee)
        self.shift_remove_btn.clicked.connect(self.removeShifts)
        self.shift_start_btn.clicked.connect(self.startShift)
        self.shift_stop_btn.clicked.connect(self.finishShift)
        self.shifts_date_filter_dateEdit1.dateChanged['QDate'].connect(lambda date : self.shifts_sort_model.setDateFilter(date))
        self.shifts_clear_btn.clicked.connect(self.clearShiftsFitlers)

        # Reports tab buttons
        self.report_export_btn.clicked.connect(lambda : self.exportTable(self.reports_model))
        self.report_search_btn.clicked.connect(lambda : self.toggleMenuMaxWidth(self.date_treeview_panel, 300, True))
        self.report_search_type_comboBox.currentTextChanged['QString'].connect(self.reportType)
        self.report_remove_btn.clicked.connect(self.removeReport)
        self.report_clear_btn.clicked.connect(self.clearReportsFitlers)

        # Offers tab buttons
        self.offer_add_btn2.clicked.connect(lambda: self.toggleMenuMaxWidth(self.frame_33,500,True))
        self.offer_add_btn.clicked.connect(self.addOffer)
        self.offer_remove_btn.clicked.connect(self.removeOffer)
        self.plus_item_btn.clicked.connect(self.plusOffer)
        self.offers_item_name_filter_txt.textChanged['QString'].connect(lambda offer_name : self.offers_sort_model.setItemNameFilter(offer_name.strip()))
        self.offers_date_filter_dateEdit1.dateChanged['QDate'].connect(lambda date : self.offers_sort_model.setDateFilter(date))
        self.offers_clear_btn.clicked.connect(self.clearOffersFitlers)

        # Copy button
        self.copy_delete_btn.clicked.connect(self.copyAndDelete)

    ##################
    # Toggling Forms #
    ##################
    def toggleStackWidget(self, maxWidth, enable):
        if enable:
    
            # GET WIDTH
            width = self.stackedWidget.width()

            # SET MAX WIDTH
            widthExtended = maxWidth


            # width ANIMATION
            self.animation = QtCore.QPropertyAnimation(self.stackedWidget, b"maximumWidth")
            self.animation.setDuration(250)
            self.animation.setStartValue(width)
            self.animation.setEndValue(widthExtended)
            self.animation.setEasingCurve(QtCore.QEasingCurve.OutQuad)

            self.animation.start()

    def toggleMainButtonsMenu(self, frame, maxWidth, enable):
        if enable:
    
            # GET WIDTH
            width = frame.width()
            standard_width = 70

            # SET MAX WIDTH
            if (width == 70):
                widthExtended = maxWidth

            elif(maxWidth == 0):
                widthExtended = 0

            else:
                widthExtended = standard_width
            
            # width ANIMATION
            self.animation = QtCore.QPropertyAnimation(frame, b"maximumWidth")
            self.animation.setDuration(250)
            self.animation.setStartValue(width)
            self.animation.setEndValue(widthExtended)
            self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)

            self.group = QtCore.QSequentialAnimationGroup(self)
            self.group.addAnimation(self.animation)
            self.group.start()
            
    def toggleMenuMaxWidth(self, frame, maxWidth, enable):
        if enable:
    
            # GET WIDTH
            width = frame.width()
            standard_width = 0

            # SET MAX WIDTH
            if width == 0:
                widthExtended = maxWidth

            else:
                widthExtended = standard_width

            # width ANIMATION
            self.animation = QtCore.QPropertyAnimation(frame, b"maximumWidth")
            self.animation.setDuration(250)
            self.animation.setStartValue(width)
            self.animation.setEndValue(widthExtended)
            self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuart)
            self.animation.start()
            
    def toggleMenuMaxHeight(self, frame, maxHeight, enable):
        if enable:
            standard_height = 0
    
            # GET WIDTH
            height = frame.height()

            # SET MAX WIDTH
            if height == 100:
                heightExtended = maxHeight
            
            elif(maxHeight == 0):
                heightExtended = 0

            else:
                heightExtended = standard_height
            
           
            #  height ANIMATION
            self.animation = QtCore.QPropertyAnimation(frame, b"maximumHeight")
            self.animation.setDuration(250)
            self.animation.setStartValue(height)
            self.animation.setEndValue(heightExtended)
            self.animation.setEasingCurve(QtCore.QEasingCurve.InOutQuad)
            self.group = QtCore.QSequentialAnimationGroup(self)
            self.group.addAnimation(self.animation)
            self.group.start()

    def slideErrorFrame(self, maxHeight, error_msg, enable):
        if enable:
            # set error msg in error label            
            self.error_lbl.setText(error_msg)

            # GET WIDTH
            height = self.error_frame.height()

            # SET MAX WIDTH
            heightExtended = maxHeight

            # width ANIMATION
            self.animation = QtCore.QPropertyAnimation(self.error_frame, b"maximumHeight")
            self.animation.setDuration(250)
            self.animation.setStartValue(height)
            self.animation.setEndValue(heightExtended)
            self.animation.setEasingCurve(QtCore.QEasingCurve.OutQuad)

            self.group = QtCore.QSequentialAnimationGroup(self)
            self.group.addAnimation(self.animation)
            self.group.start()


    ##########
    # BackToMain # 
    ##########
    def BackToMain(self):
        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'Main_tab'))
        self.toggleMainButtonsMenu(self.main_buttons_frame, 0, True)
        self.toggleStackWidget(0,True)

        self.frame_19.setMaximumWidth(0)
        self.frame_11.setMaximumWidth(0)
        self.frame_21.setMaximumWidth(0)
        self.frame_24.setMaximumWidth(0)
        self.date_treeview_panel.setMaximumWidth(0)

        self.panel_title_lbl.setText('')

        self.blur.setBlurRadius(1)

        # Reset the FLAGS
        self._DAILY_TABLES_FLAG = False
        self._ARCHIVE_TABLES_FLAG = False
        self._SETTINGS_TABLES_FLAG = False

    ###################
    # Copy and Delete #
    ###################
    def copyAndDelete(self):
        messageBox = QtWidgets.QMessageBox.information(
                self,
                "Copy data to archive",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            )
        if messageBox == QtWidgets.QMessageBox.Ok:
            updateReports(db= self.daily_conn)
            copyData(src_db_1 = self.daily_conn, dest_db_2 = self.archive_conn)

    ##########
    # Panels #
    ##########
    def dailyPanel(self):
        """Open daily panel"""
        
        # Check if any shift is started or not
        if(not self.isAnyShiftActive()):
            return

        self._DAILY_TABLES_FLAG = True
        self.toggleMainButtonsMenu(self.main_buttons_frame, 70, True)
        self.panel_title_lbl.setText('Daily')
        
        # Change blur radius for main panel
        self.blur.setBlurRadius(5)

        # Disable main buttons after clicking on any of them
        self.disableMainButton()
        
        # Disable search buttons
        self.daily_customer_search_btn.setEnabled(False)
        self.monthly_customer_search_btn.setEnabled(False)
        self.order_search_btn.setEnabled(False)
        

        # Enable Add buttons
        self.daily_customer_add_btn2.setEnabled(True)
        self.daily_customer_add_order.setEnabled(True)
        self.monthly_customer_add_btn2.setEnabled(True)
        self.monthly_customer_update_btn.setEnabled(True)
        self.order_add_btn2.setEnabled(True)
        
        self.buttons_stackedWidget.setCurrentWidget(self.buttons_stackedWidget.findChild(QtWidgets.QWidget, 'dailyAndarchive_buttons_tab'))
        
        # Reinitialize model objects by daily models
        self.setViewModel()

        # Reinitialize comboBox
        self.initialComboBoxs(self.daily_conn)

    def archivePanel(self):
        """Open acrchive panel"""

        # hide shift error frame
        self.slideErrorFrame(0, '', True)

        self._ARCHIVE_TABLES_FLAG = True
        self.toggleMainButtonsMenu(self.main_buttons_frame, 70, True)
        self.panel_title_lbl.setText('Archive')  

        self.buttons_stackedWidget.setCurrentWidget(self.buttons_stackedWidget.findChild(QtWidgets.QWidget, 'dailyAndarchive_buttons_tab'))

        # Change blur radius for main panel
        self.blur.setBlurRadius(5)

        # Disable main buttons after clicking on any of them
        self.disableMainButton()

        # Enable search buttons
        self.daily_customer_search_btn.setEnabled(True)
        self.monthly_customer_search_btn.setEnabled(True)
        self.order_search_btn.setEnabled(True)

        # Disable buttons
        self.daily_customer_add_btn2.setEnabled(False)
        self.daily_customer_add_order.setEnabled(False)
        self.monthly_customer_add_btn2.setEnabled(False)
        self.monthly_customer_update_btn.setEnabled(False)
        self.order_add_btn2.setEnabled(False)
        


        # Reinitialize model objects by daily models
        self.setViewModel()
    
        # Reinitialize comboBox
        self.initialComboBoxs(self.archive_conn)
    
    def settingsPanel(self):
        """Open settings panel"""

        # hide shift error frame
        self.slideErrorFrame(0, '', True)
        
        
        self._SETTINGS_TABLES_FLAG = True
        self.toggleMainButtonsMenu(self.main_buttons_frame, 70, True)
        self.panel_title_lbl.setText('Settings')  
        
        self.buttons_stackedWidget.setCurrentWidget(self.buttons_stackedWidget.findChild(QtWidgets.QWidget, 'settings_buttons_tab'))
        
        # Change blur radius for main panel
        self.blur.setBlurRadius(5)

        # Disable main buttons after clicking on any of them
        self.disableMainButton()


        # Reinitialize model objects by daily models
        self.setViewModel()

        # Reinitialize comboBox
        self.initialSettingsComboBoxs()

    def setViewModel(self):
        
        if(self._DAILY_TABLES_FLAG == True):
            self.daily_customers_model = DailyCustomersModel(self.daily_conn, self)
            self.monthly_customers_model = MonthlyCustomersModel(self.daily_conn)
            self.orders_model = OrdersModel(self.daily_conn)

            self.daily_customers_sort_model = DailyCustomersSortModel(self.daily_customers_model)
            self.monthly_customers_sort_model = MonthlyCustomersSortModel(self.monthly_customers_model)
            self.orders_sort_model = OrdersSortModel(self.orders_model)

        elif(self._ARCHIVE_TABLES_FLAG == True):
            self.daily_customers_model = DailyCustomersModel(self.archive_conn)
            self.monthly_customers_model = MonthlyCustomersModel(self.archive_conn)
            self.orders_model = OrdersModel(self.archive_conn)

            self.daily_customers_sort_model = DailyCustomersSortModel(self.daily_customers_model)
            self.monthly_customers_sort_model = MonthlyCustomersSortModel(self.monthly_customers_model)
            self.orders_sort_model = OrdersSortModel(self.orders_model)

        elif(self._SETTINGS_TABLES_FLAG == True):
            self.warehouse_model = WarehouseModel(self.daily_conn)
            self.employees_model = EmployeesModel(self.daily_conn)
            self.shifts_model = ShiftsModel(self.daily_conn)
            self.offers_model = OffersModel(self.daily_conn)
            self.reports_model = ReportsModel(self.daily_conn)

            self.warehouse_sort_model = WarehouseSortModel(self.warehouse_model)
            self.employees_sort_model = EmployeesSortModel(self.employees_model)
            self.shifts_sort_model = ShiftsSortModel(self.shifts_model)
            self.offers_sort_model = OffersSortModel(self.offers_model)
            self.reports_sort_model = ReportsSortModel(self.reports_model)

            self.setCurrentDate()

    def disableMainButton(self):
        self.archive_btn.setEnabled(False)
        self.daily_btn.setEnabled(False)
        self.settings_btn.setEnabled(False)
        self.exit_btn.setEnabled(False)
    
    def isAnyShiftActive(self, exit_event=False):
        ret = checkShiftActive(db = self.daily_conn)
        if(ret == 1):
            if(exit_event == False):
                self.slideErrorFrame(0, '', True)
            return True

        elif(ret == 0):
            if(exit_event == False):
                self.slideErrorFrame(100,'Please start a new shift', True)
            return False
        
        

    ###################
    # Daily customers #
    ###################
    def showDailyCustomers(self):
        """Show all daily customers from daily table"""
        
        # mapper = QtWidgets.QDataWidgetMapper(self)
        # mapper.setOrientation(Qt.Horizontal)
        # mapper.setModel(self.daily_customers_model.model)

        # mapper.addMapping(self.daily_customer_name_txt, 1)
        # mapper.addMapping(self.daily_customer_monthID_txt, 2)
        # mapper.toFirst()

        self.setViewModel()

        if(self._ARCHIVE_TABLES_FLAG):
            self.showYMDs(self.archive_conn, 'Daily_customers', 'daily_date')

        self.daily_customers_tableView.setModel(self.daily_customers_sort_model)
        self.daily_customers_tableView.setItemDelegate(QSqlRelationalDelegate(self.daily_customers_tableView))
        # self.daily_customers_tableView.hideColumn(0)

        # Set on Daily customers tab widget
        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'daily_customers_tab'))
        self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'daily_customers_properties_panel'))
        self.toggleStackWidget(75, True)

        self.daily_customers_tableView.selectionModel().selectionChanged.connect(lambda : self.daily_customers_count_lbl.setText(str(len(self.daily_customers_tableView.selectionModel().selectedRows()))))
        
    def addDailyCustomer(self):
        """Add new daily customer to daily table"""

        if(not self.isAnyFeeZero()):
            return

        # Check if the customer name field is empty
        if(self.daily_customer_name_txt.text()==""):
            self.daily_customer_name_txt.setFocus()
            QtWidgets.QToolTip.showText(self.daily_customer_name_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter customer name")
            return
        
            
        data = [self.daily_customer_name_txt.text().strip(),self.daily_customer_cusType_comboBox.currentText(), self.daily_customer_monthID_txt.text().strip()]

        ret = self.daily_customers_model.addDailyCustomer(data)

        if(ret):
            # # scroll to inserted row
            # flags = QtCore.QItemSelectionModel.ClearAndSelect | QtCore.QItemSelectionModel.Rows
            # print(self.daily_customers_model.query().lastInsertId())
            # print(type(self.daily_customers_model.query().lastInsertId)())
            # index = self.daily_customers_tableView.model().index(self.daily_customers_model.query().lastInsertId(), 0)       
            # self.daily_customers_tableView.scrollTo(index)
            # # self.daily_customers_tableView.scrollToBottom()
            # self.daily_customers_tableView.selectionModel().select(index, flags)

            # Reset texts
            self.daily_customer_name_txt.setText('')
            self.daily_customer_monthID_txt.setText('')
            self.daily_customer_cusType_comboBox.setCurrentText('')

            # Renew the completers
            self.Completers()

            # Update Reports
            updateReports(db =self.daily_conn)

            # Reinitial the comboBoxs
            if(self._DAILY_TABLES_FLAG == True):
                self.initialComboBoxs(self.daily_conn)
            elif(self._ARCHIVE_TABLES_FLAG == True):
                self.initialComboBoxs(self.archive_conn)

        else:
            QtWidgets.QMessageBox.warning(
            self,
            "Info",
            "Record exists...!",
            QtWidgets.QMessageBox.Ok ,
            )
            
    def removeDailyCustomer(self):
        """Remove customer from table"""
        # row = self.daily_customers_tableView.currentIndex().row()

        # Get selected rows and their indexs
        rows_indices = self.daily_customers_tableView.selectionModel().selectedRows()
        
        if (len(rows_indices) < 1):
            QtWidgets.QToolTip.showText(self.daily_customer_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select customer/s")
    
        else:    
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
        
            
                # Create a worker thread
                if(self._DAILY_TABLES_FLAG):
                    con = self.daily_conn
                elif(self._ARCHIVE_TABLES_FLAG):
                    con = self.archive_conn

                worker = Worker(con, 'Daily_customers', 'daily_id', rows_indices)
                
                # Disable remove button
                self.daily_customer_remove_btn.setEnabled(False)

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.daily_customer_remove_btn.setEnabled(True))
                worker.finished.connect(lambda : self.daily_customers_model.submitAll())
                worker.finished.connect(lambda : self.daily_customers_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Daily_customers', column = 'daily_id', db1 = self.daily_conn ,db2 = self.archive_conn))
                worker.finished.connect(lambda : updateReports(db =self.daily_conn))
                worker.finished.connect(lambda : self.initialComboBoxs(con))

                
                
                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()

                

            # Renew the completers
            self.Completers()

    @QtCore.pyqtSlot(str)
    def setDailyName(self, id):
        """Set customer name by enter his/her monthly id"""
        if(id != ''):
            name = retrieveMonthlyNames(id = id, db  = self.daily_conn)
            if(name != []):
                self.daily_customer_name_txt.setText(name[0])
    
    @QtCore.pyqtSlot(str)
    def setDailyId(self, name):
        """Set customer monthly id by enter his/her name"""
        if(name != ''):
            id = retrieveMonthlyid(name = str(name).strip(), db  = self.daily_conn)
            if(id != None):
                self.daily_customer_monthID_txt.setText(str(id))
            
            elif( id == None):
                self.daily_customer_monthID_txt.setText('')
    
    def addOrderDirectly(self):
        """Add order directly by selecting customer from daily table"""
        row = self.daily_customers_tableView.currentIndex().row()
        if (row < 0):
            QtWidgets.QToolTip.showText(self.daily_customer_add_order.mapToGlobal(QtCore.QPoint(0,30)),"Select customer/s")
        
        # If customer is selected 
        else:
            # Get the customer name
            daily_name = self.daily_customers_tableView.model().data(self.daily_customers_tableView.model().index(row,1), Qt.EditRole)
            
            self.customer_name_txt2.setText(daily_name)
            
            # Move to orders tab
            self.showOrders()
            self.stackedWidget.setCurrentIndex(1)

            if(self.frame_11.width() == 0):
                self.toggleMenuMaxWidth(self.frame_11,500,True)
        # self.daily_customer_monthID_txt.setText(retrievMonthlyid(name = self.daily_customer_name_txt.text()))

    def clearDailyFilters(self):
        self.daily_customer_name_filter_txt.setText('')
        self.daily_customer_subsType_filter_comboBox.setCurrentText('')
        self.daily_customers_sort_model.setDateFilter('')


    #####################
    # Monthly customers #
    #####################
    def showMonthlyCustomers(self):
        """Show all monthly customers from Monthly table"""
        self.setViewModel()

        if(self._ARCHIVE_TABLES_FLAG):
            self.showYMDs(self.archive_conn, 'Monthly_customers', 'start_date')

        self.monthly_customers_tableView.setModel(self.monthly_customers_sort_model)

        # Set on Monthly customers tab widget
        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'monthly_tab'))
        self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'monthly_customers_properties_panel'))
        self.toggleStackWidget(75, True)

        self.monthly_customers_tableView.selectionModel().selectionChanged.connect(lambda : self.monthly_customers_count_lbl.setText(str(len(self.monthly_customers_tableView.selectionModel().selectedRows()))))

    def addMonthlyCustomer(self):
        """Add new monthly customer to Monthly table"""

        if(not self.isAnyFeeZero()):
            return

        # Check if the customer name field is empty
        if(self.monthly_customer_name_txt.text()==""):
            self.monthly_customer_name_txt.setFocus()
            QtWidgets.QToolTip.showText(self.monthly_customer_name_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter customer name")
            return

        # Check if the customer subscription field is empty
        if(self.monthly_customer_subscritption_comboBox.currentText()==""):
            self.monthly_customer_subscritption_comboBox.setFocus()
            QtWidgets.QToolTip.showText(self.monthly_customer_subscritption_comboBox.mapToGlobal(QtCore.QPoint(0,30)),"Select subsription type")
            return
        
        data = [self.monthly_customer_name_txt.text().strip(), self.monthly_customer_subscritption_comboBox.currentText().strip()]
        
        ret = self.monthly_customers_model.addMonthlyCustomer(data)

        if(ret):
            # Reset text
            self.monthly_customer_name_txt.setText('')
            self.monthly_customer_subscritption_comboBox.setCurrentText('')

            # Renew the completers
            self.Completers()

            # Update Reports
            updateReports(db =self.daily_conn)

            # Reinitial the comboBoxs
            if(self._DAILY_TABLES_FLAG == True):
                self.initialComboBoxs(self.daily_conn)
            elif(self._ARCHIVE_TABLES_FLAG == True):
                self.initialComboBoxs(self.archive_conn)


        else:
            QtWidgets.QMessageBox.warning(
            self,
            "Info",
            "Record exists...!",
            QtWidgets.QMessageBox.Ok ,
            )    
          
    def removeMonthlyCustomer(self):
        """Remove monthly customer from Monthly table""" 

        # Get selected rows and their indexs
        rows_indices = self.monthly_customers_tableView.selectionModel().selectedRows()
        
        
        if (len(rows_indices) < 1) :
            QtWidgets.QToolTip.showText(self.monthly_customer_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select customer/s/s")

        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                

                # Create a worker thread
                if(self._DAILY_TABLES_FLAG):
                    con = self.daily_conn
                elif(self._ARCHIVE_TABLES_FLAG):
                    con = self.archive_conn

                worker = Worker(con, 'Monthly_customers', 'monthly_id', rows_indices)

                # Disable remove button
                self.monthly_customer_remove_btn.setEnabled(False)

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.monthly_customer_remove_btn.setEnabled(True))
                worker.finished.connect(lambda : self.monthly_customers_model.submitAll())
                worker.finished.connect(lambda : self.monthly_customers_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Monthly_customers', column = 'monthly_id', db1 = self.daily_conn))
                worker.finished.connect(lambda : updateReports(db =self.daily_conn))
                worker.finished.connect(lambda : self.initialComboBoxs(con))

                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()
          
    def updateMonthlyCustomer(self):
        """Update subscription state for selected Monthly customer""" 
        # Get selected rows and their indexs
        rows_indices = self.monthly_customers_tableView.selectionModel().selectedRows()

        if (len(rows_indices) <= 0):
            QtWidgets.QToolTip.showText(self.monthly_customer_update_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select customer/s")

        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Renew monthly subscription",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel
            )
            if messageBox == QtWidgets.QMessageBox.Ok:

                for index in rows_indices:
                    monthly_id = self.monthly_customers_tableView.model().data(self.monthly_customers_tableView.model().index(index.row(),0),Qt.EditRole)
                    updateSubsState(monthly_id, self.daily_conn)
            
            
            self.showMonthlyCustomers()

    def clearMonthlyFilters(self):

        self.monthly_customer_name_filter_txt.setText('')
        self.monthly_customer_subsState_filter_comboBox.setCurrentText('')
        self.monthly_customer_subsType_filter_comboBox.setCurrentText('')

        self.monthly_customers_sort_model.setDateFilter('')


    ##########
    # Orders #
    ##########
    def showOrders(self):
        """Show all Orders from Orders table"""  
        self.setViewModel()

        if(self._ARCHIVE_TABLES_FLAG):
            self.showYMDs(self.archive_conn, 'Orders', 'order_date')


        self.orders_tableView.setModel(self.orders_sort_model)
        self.orders_tableView.setItemDelegate(QSqlRelationalDelegate(self.orders_tableView))
        self.orders_tableView.hideColumn(0)
        
        # Set Orders tab widget
        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'orders_tab'))
        self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'orders_properties_panel'))
        self.toggleStackWidget(75, True)
        
        self.orders_tableView.selectionModel().selectionChanged.connect(lambda : self.orders_count_lbl.setText(str(len(self.orders_tableView.selectionModel().selectedRows()))))

    def addOrder(self):

        """Add new order to Orders table"""
        data = []

        # Check if the customer name field is empty
        if(self.order_sell_type_comboBox.currentText()=='عام' and self.customer_name_txt2.text()==""):
            self.customer_name_txt2.setFocus()
            QtWidgets.QToolTip.showText(self.customer_name_txt2.mapToGlobal(QtCore.QPoint(0,30)),"Enter customer name")
            return
        
        # Check if the customer name field is empty
        elif(self.order_sell_type_comboBox.currentText()=='ضيافة' and self.customer_name_txt2.text()!=""):
            self.customer_name_txt2.setFocus()
            QtWidgets.QToolTip.showText(self.customer_name_txt2.mapToGlobal(QtCore.QPoint(0,30)),"Delete the customer name")
            return

        # Take children of order register panel 
        # Item and quantity
        order_comboBox_lst = self.orders_items_frame.findChildren(QtWidgets.QComboBox)
        order_quantity_txt_lst = self.orders_items_frame.findChildren(QtWidgets.QLineEdit)

        # Split quantity texts from other texts
        order_quantity_txt_lst = [order_quantity_txt_lst[i] for i in range(len(order_quantity_txt_lst))if i%2 != 0]
       
        # Iterate to check if any Item name or quantity is empty
        for order_comboBox,quantity_txt in zip(order_comboBox_lst,order_quantity_txt_lst):
            # Check if the item comboBox is empty
            if(order_comboBox.currentText()==""):
                order_comboBox.setFocus()
                QtWidgets.QToolTip.showText(order_comboBox.mapToGlobal(QtCore.QPoint(0,30)),"Select item")
                return
            # Check if the quantity field is empty
            elif(quantity_txt.text()==""):
                quantity_txt.setFocus()
                QtWidgets.QToolTip.showText(quantity_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter quantity")
                return      

        # Iterate to add customer's orders_model
        items_data = [(order_comboBox.currentText(), quantity_txt.text()) for order_comboBox, quantity_txt in zip(order_comboBox_lst, order_quantity_txt_lst)]
        
        data = [(self.customer_name_txt2.text(), self.order_sell_type_comboBox.currentText(), self.offer_id), items_data, int(self.total_price_lbl.text())]

        ret = self.orders_model.addOrder(data)

        if(self.orders_model.lastError().text().startswith("exceed")):
            QtWidgets.QMessageBox.critical(
            self,
            "ِAlert",
            "Exceed available quantity",
            QtWidgets.QMessageBox.Ok,
            )
            
            return
    
        if(ret):
            # Delete order registring panel 
            self.clearLayout(self.orders_items_frame.layout())

            # Resest orders_number variable
            self._order_item_number = 0
            self.offer_id = None

            # Insert an empty order
            self.plusOrder()

            # Update Reports
            updateReports(db =self.daily_conn)

            # Reinitial the comboBoxs
            if(self._DAILY_TABLES_FLAG == True):
                self.initialComboBoxs(self.daily_conn)
            elif(self._ARCHIVE_TABLES_FLAG == True):
                self.initialComboBoxs(self.archive_conn)

            # Reset total price label and customer name
            self.total_price_lbl.setText(str(0))
            self.customer_name_txt2.setText('')
 
    def removeOrder(self):
        """Remove order from Orders table"""
        rows_indices = self.orders_tableView.selectionModel().selectedRows()
        
        if (len(rows_indices) < 1):
            QtWidgets.QToolTip.showText(self.order_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select order/s")

        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                
            
                # Create a worker thread
                if(self._DAILY_TABLES_FLAG):
                    con = self.daily_conn
                elif(self._ARCHIVE_TABLES_FLAG):
                    con = self.archive_conn

                worker = Worker(con, 'Orders', 'order_id', rows_indices)

                # Disable remove button
                self.order_remove_btn.setEnabled(False)

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.order_remove_btn.setEnabled(True))
                worker.finished.connect(lambda : self.orders_model.submitAll())
                worker.finished.connect(lambda : self.orders_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Orders', column = 'order_id', db1 = self.daily_conn))
                worker.finished.connect(lambda : updateReports(db =self.daily_conn))
                worker.finished.connect(lambda : self.initialComboBoxs(con))

                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()
            
    def plusOrder(self):
        """Adding new order to orders list"""
        self._order_item_number += 1
        
        order_frame = CustomAddOrderFrame(self._order_item_number, self.orders_items_frame)

        # Add the container frame into parent frame layout
        self.verticalLayout_70.addWidget(order_frame)

        
        # pass on items comboBox
        order_frame.order_item_comboBox.addItems(['']+retrieveItemNames(name_filter=('',''), db = self.daily_conn))

        # self.changeComboItems(order_frame.order_item_comboBox)
        # order_frame.order_item_comboBox.currentTextChanged['QString'].connect(lambda : self.changeComboItems2(order_frame.order_item_comboBox))

        # Connect each comboBox with its label.
        order_frame.order_item_comboBox.currentTextChanged['QString'].connect(lambda : self.setOrderItemPrice(order_frame.order_item_comboBox, order_frame.order_item_price_lbl, order_frame.order_quantity_txt))
        order_frame.order_item_comboBox.currentTextChanged['QString'].connect(lambda : self.isOrderItemExist(order_frame.order_item_comboBox, order_frame.order_quantity_txt))
       
        # pass on items comboBox to connect it with its label.
        order_frame.order_quantity_txt.textEdited['QString'].connect(lambda : self.checkOrderItemQauntity(order_frame.order_item_comboBox, order_frame.order_quantity_txt))
        
        
        # pass on delete buttons to connect it with its frame.
        order_frame.delete_btn.clicked.connect(self.checkOfferAvailable)
        order_frame.delete_btn.clicked.connect(lambda : self.deleteOrderFrame(order_frame))

    def changeComboItems(self, order_item_comboBox):
        # Check if the item is selected previously 

        combo_selected_items = [item.currentText() for item in self.orders_items_frame.findChildren(QtWidgets.QComboBox)]
        combo_selected_items.append('')
        combo_selected_items = tuple(combo_selected_items)

        # order_item_comboBox = self.orders_items_frame.findChildren(QtWidgets.QComboBox)[-1]
        # order_item_comboBox.clear()
        order_item_comboBox.addItem('')
        order_item_comboBox.addItems(retrieveItemNames(name_filter=combo_selected_items, db = self.daily_conn))      
    
    def changeComboItems2(self, order_item_comboBox):

        # Change items in other comboBoxs
        combo_selected_items = [item for item in self.orders_items_frame.findChildren(QtWidgets.QComboBox)]
        combo_selected_items[combo_selected_items.index(order_item_comboBox)+1:]

        for combo in combo_selected_items:
            combo.clear()
            combo.addItem('')
            combo.addItems(retrieveItemNames(name_filter=(order_item_comboBox.currentText(),''), db = self.daily_conn))      

    def isOrderItemExist(self, selected_item, quntity_txtBox):

        # Check if the item is selected previously 
        selected_items_name = [item.currentText() for item in self.orders_items_frame.findChildren(QtWidgets.QComboBox)]

        if(retrieveItemId(selected_item.currentText(), db = self.daily_conn) is None):
            QtWidgets.QToolTip.showText(selected_item.mapToGlobal(QtCore.QPoint(0,30)),"Not found")
            selected_item.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 255, 0);\n"
            "}"
            )
            quntity_txtBox.setEnabled(False)

        elif(selected_item.currentText()==''):
            selected_item.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 170, 0);\n"
            "}"
            )
            quntity_txtBox.setEnabled(False)

        # Check if the item selected previously
        elif(selected_items_name.count(selected_item.currentText()) > 1):
            QtWidgets.QToolTip.showText(selected_item.mapToGlobal(QtCore.QPoint(0,30)),"Selected previously")
            selected_item.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 0, 0);\n"
            "}"
            )

            # Disable editing quantity text 
            quntity_txtBox.setEnabled(False)
            quntity_txtBox.setText('')
        
        else:
            self.checkOfferAvailable()
 
    def setOrderItemPrice(self, combo_selected_item, price_lbl, quntity_txt):
        """Get the item price and add it to label"""
            
        price = retrieveItemPrice(combo_selected_item.currentText(), self.daily_conn)
        price_lbl.setText(str(price)+' L.S')
        combo_selected_item.setStyleSheet(
        "QComboBox{\n"
        "border-width:0px 0px 4px 0px;\n"
        "border-style: solid;\n"
        "border-radius:0px;\n"
        "border-color: rgb(255, 170, 0);\n"
        "}"
        )

        # Allow editing quantity text 
        quntity_txt.setEnabled(True)

    def checkOrderItemQauntity(self, combo_selected_item, quantity_txt):
        if(quantity_txt.text() ==''):
            return

        remain_quantity = retrieveItemAvailabelQuantity(retrieveItemId(combo_selected_item.currentText(), db = self.daily_conn), db = self.daily_conn)

        # If not enough items exist 
        if(int(quantity_txt.text()) > remain_quantity):
            QtWidgets.QToolTip.showText(quantity_txt.mapToGlobal(QtCore.QPoint(0,30)),f"Exceed {remain_quantity}")
            
            quantity_txt.setStyleSheet("QLineEdit{\n"
            "\n"
            "    border-style: solid;\n"
            "    border-width: 4px 4px 4px 4px;\n"
            "    border-radius: 0px;    \n"
            "    border-color: rgb(255, 0, 0);\n"
            "}\n"
            "")
            self.order_add_btn.setEnabled(False)

        elif(int(quantity_txt.text()) <= remain_quantity):
            quantity_txt.setStyleSheet("QLineEdit{\n"
            "\n"
            "    border-style: solid;\n"
            "    border-width: 0px 0px 4px 0px;\n"
            "    border-radius: 0px;    \n"
            "    border-color: rgb(255, 170, 0);\n"
            "}\n"
            "")
            self.order_add_btn.setEnabled(True)

            self.checkOfferAvailable()

    def checkOfferAvailable(self): 

        # Check if the selected items have an offer
        selected_items_name = [item.currentText() for item in self.orders_items_frame.findChildren(QtWidgets.QComboBox)]
        selected_items_ids = [retrieveItemId(item_name, self.daily_conn) for item_name in selected_items_name]
        
        quantities = self.orders_items_frame.findChildren(QtWidgets.QLineEdit)

        # Check if all quantities text box are enabled
        enabled = list(map(lambda x : x.isEnabled(), quantities))
        if(False in enabled):
            self.offer_id = None
            return
            
        entered_quantities = [quantities[i].text() for i in range(len(quantities)) if (i%2 != 0)]
        entered_quantities = [int(quantity) for quantity in entered_quantities if(quantity.isdigit())]

        offers_items = retrieveOffersItems(with_date= False, with_quantity=True, db = self.daily_conn).items()
        for offer_key, items_key in offers_items:
            
            # get offer's items quantities
            quan = [item for sublist in list(items_key.values()) for item in sublist]

            if(list(items_key.keys()) == sorted(selected_items_ids)):
                if(quan == entered_quantities):
                    self.offer_id = offer_key
                        
                    for item in self.orders_items_frame.findChildren(QtWidgets.QComboBox):
                        item.setStyleSheet(
                        "QComboBox{\n"
                        "border-width:0px 0px 4px 0px;\n"
                        "border-style: solid;\n"
                        "border-radius:0px;\n"
                        "border-color: rgb(0, 255, 0);\n"
                        "}"
                        )
                    
            else:
                self.offer_id = None
                for item in self.orders_items_frame.findChildren(QtWidgets.QComboBox):
                    item.setStyleSheet(
                    "QComboBox{\n"
                    "border-width:0px 0px 4px 0px;\n"
                    "border-style: solid;\n"
                    "border-radius:0px;\n"
                    "border-color: rgb(255, 170, 0);\n"
                    "}"
                    )

            # Change the total price
            self.setTotalOrderPrice(self.offer_id)


        if(not offers_items):
            self.setTotalOrderPrice(self.offer_id)

    def setTotalOrderPrice(self, offer_id = None):
        """Calculate the items selected total price"""
        
        if(self.order_sell_type_comboBox.currentText()=='ضيافة'):
            self.total_price_lbl.setText('0')
            return

        # Take children of order register panel (Items and quantities)
        selected_items_name = self.orders_items_frame.findChildren(QtWidgets.QComboBox)
        quantities = self.orders_items_frame.findChildren(QtWidgets.QLineEdit)

        # Split quantity texts from other texts
        quantities = [quantities[i].text() for i in range(len(quantities))if i%2 != 0]
        
        total = 0

        if(offer_id is None):
            # Iterate to sum the items price
            for item, quantity in zip(selected_items_name, quantities):
                price = retrieveItemPrice(item.currentText(), self.daily_conn)
                if(quantity.isdigit()):
                    quantity = int(quantity)
                    total += price * quantity

        elif(offer_id is not None):
            total = retrieveOfferPrice(offer_id, db = self.daily_conn)
        
        # Put total price in label
        self.total_price_lbl.setText(f'{total}')

    def clearOrderFitlers(self):
    
        self.orders_customer_name_filter_txt.setText('')
        self.orders_type_filter_comboBox.setCurrentText('')
        self.orders_sort_model.setDateFilter('')

    def deleteOrderFrame(self, frame):
        """Delete frame after clicking"""

        # Keep at least one order that shouldn't be deleted
        if len(self.orders_items_frame.findChildren(QtWidgets.QComboBox)) > 1:

            sip.delete(frame) 

            # decrement order number
            self._order_item_number -=1

            # Reset item labels name
            _translate = QtCore.QCoreApplication.translate
            labels = [label for label in self.orders_items_frame.findChildren(QtWidgets.QLabel) if(label.objectName() == 'item_number_lbl')]
            for num, label in zip(range(1, self._order_item_number+1),labels): 
                label.setText(_translate("MainWindow", f"Item : {num} "))
            
            # Recalculate the total order price
            self.checkOfferAvailable()
 
    def clearLayout(self, layout):
        """Clear frame layout after insert new order to tabel"""
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clearLayout(item.layout())
            sip.delete(widget)
   
    def resetAvailableItemsComboBox(self):
        """Refresh items comboBox list after adding new item to Warehouse"""
        # pass on items comboBox
        comboBox_lst = self.orders_items_frame.findChildren(QtWidgets.QComboBox)

        for comboBox in comboBox_lst:
            # Add Warehouse item to each comboBox
            comboBox.clear()
            comboBox.addItem('')
            comboBox.addItems(retrieveItemNames(db  = self.daily_conn))
    
    def openOrderDialog(self):
        """Open change order dialog"""
        # Get selected row and its index
        index = self.orders_tableView.selectionModel().currentIndex()

        order_id = index.sibling(index.row(), self.orders_model.fieldIndex('order_id')).data()
        customer_name = index.sibling(index.row(), self.orders_model.fieldIndex('daily_name')).data()
        order_type = index.sibling(index.row(), self.orders_model.fieldIndex('order_type')).data()

        if (order_id == None):
            QtWidgets.QToolTip.showText(self.order_show_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select order")
        
        else:
                order_diag = OrderDialog(order_id, customer_name, order_type, db = self.daily_conn)
                if(order_diag.exec() == QtWidgets.QDialog.Accepted):
                    pass  
    
    #############
    # Warehouse #
    #############
    def showWarehouse(self):        
        """Show all items from Warehouse table"""

        # If Daily panel is opend then we can make change to date (add, edit, ...)
        if(self._SETTINGS_TABLES_FLAG == True):
            self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'warehouse_properties_panel'))
            self.toggleStackWidget(75, True)


        self.warehouse_tableView.setModel(self.warehouse_sort_model)

        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'warehouse_tab'))
        self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'warehouse_properties_panel'))
        
        # Read available item type 
        # self.initialItemsTypeComboBox()

    def addItem(self):
        """Add new itme to Warehouse table"""
        # Check if the item name field is empty
        if(not self.warehouse_item_name_txt.text()):
            self.warehouse_item_name_txt.setFocus()
            QtWidgets.QToolTip.showText(self.warehouse_item_name_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter item name")
            return
        
        # Check if the item price field is empty
        if(not self.warehouse_item_price_txt.text()):
            self.warehouse_item_price_txt.setFocus()
            QtWidgets.QToolTip.showText(self.warehouse_item_price_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter price")
            return
        
        # Check if the item quantity field is empty
        if(not self.warehouse_item_quantity_txt.text()):
            self.warehouse_item_quantity_txt.setFocus()
            QtWidgets.QToolTip.showText(self.warehouse_item_quantity_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter quantity")
            return

        # Check if the item type field is empty
        if(not self.warehouse_item_type_comboBox.currentText()):
            self.warehouse_item_type_comboBox.setFocus()
            QtWidgets.QToolTip.showText(self.warehouse_item_type_comboBox.mapToGlobal(QtCore.QPoint(0,30)),"Select item type")
            return

        
        data = [
            self.warehouse_item_name_txt.text().strip(),
            int(self.warehouse_item_price_txt.text()),
            self.warehouse_item_type_comboBox.currentText().strip(), 
            int(self.warehouse_item_quantity_txt.text())
        ]
        
        ret = self.warehouse_model.addItem(data)
        

        if(ret):
            # Resest text
            self.warehouse_item_name_txt.setText('')
            self.warehouse_item_price_txt.setText('')
            self.warehouse_item_type_comboBox.setCurrentText('')
            self.warehouse_item_quantity_txt.setText('')

            # Reinialize items
            self.initialSettingsComboBoxs()

            # Renew the completers
            self.Completers()

            # Update Reports
            updateReports(db =self.daily_conn)

        else:
            QtWidgets.QMessageBox.warning(
            self,
            "Info",
            "Record exists...!",
            QtWidgets.QMessageBox.Ok ,
            )

    def removeItem(self):
        """Remove item from Warehouse table"""
        # Get selected rows and their indexs
        rows_indices = self.warehouse_tableView.selectionModel().selectedRows()
    
        if (len(rows_indices) < 1):
            QtWidgets.QToolTip.showText(self.warehouse_item_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select item/s")

        else:    
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                

                # Create a worker thread
                worker = Worker(self.daily_conn, 'Warehouse', 'item_id', rows_indices)

                # Disable remove button
                self.warehouse_item_remove_btn.setEnabled(False)

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.warehouse_item_remove_btn.setEnabled(True))
                worker.finished.connect(lambda : self.warehouse_model.submitAll())
                worker.finished.connect(lambda : self.warehouse_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Warehouse', column = 'item_id', db1 = self.daily_conn))
                worker.finished.connect(self.initialSettingsComboBoxs)
                worker.finished.connect(self.Completers)
                
                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()

    def updateItemsQauntities(self):
        messageBox = QtWidgets.QMessageBox.warning(
            self,
            "Updating current available item...alert",
            "Confirm",
            QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            QtWidgets.QMessageBox.Cancel
        )
        if (messageBox == QtWidgets.QMessageBox.Ok):
            updateCurrentItemsQuantities()


    def initialItemsTypeComboBox(self):
        self.warehouse_item_type_filter_comboBox.clear()
        self.warehouse_item_type_filter_comboBox.addItem('')
        self.warehouse_item_type_filter_comboBox.addItems(retrieveItemType(self.daily_conn))

    def clearWarehouseFitlers(self):
        self.warehouse_item_name_filter_txt.setText('')
        self.warehouse_item_type_filter_comboBox.setCurrentText('')

    ##########
    # Offers #
    ##########
    def showOffers(self):        
        """Show all items from offers table"""

        if(self._SETTINGS_TABLES_FLAG == True):
            self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'offers_properties_panel'))
            self.toggleStackWidget(75, True)


        self.offers_tableView.setModel(self.offers_sort_model)

        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'offers_tab'))

        self.showOffersItems()

        # Read available item type 
        self.initialItemsTypeComboBox()

    def showOffersItems(self):
        """Display available offers and their items"""

        self.offers_items_model = HierarcicalOffersItemsModel(self.daily_conn)
        self.offers_items_treeView.setHeaderHidden(True)
        self.offers_items_treeView.setModel(self.offers_items_model) 
        self.offers_items_treeView.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        
        # selection_model = self.offers_items_treeView.selectionModel()
        # # selection_model.selectionChanged.connect(self.selectionChangedSlot)

    def addOffer(self):
        """Add new offer with related items"""

        # Check if the offer name field is empty
        if(not self.offer_price_txt.text()):
            self.offer_price_txt.setFocus()
            QtWidgets.QToolTip.showText(self.offer_price_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter offer name")
            return
        
        # Check if the offer price field is empty
        elif(not self.offer_price_txt.text()):
            self.offer_price_txt.setFocus()
            QtWidgets.QToolTip.showText(self.offer_price_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter offer price")
            return
        
        # Take children of offer register panel 
        selected_items = self.offers_items_frame.findChildren(QtWidgets.QComboBox)
        quantities = self.offers_items_frame.findChildren(QtWidgets.QLineEdit)

        # Split quantity texts from other texts
        quantities = [quantities[i] for i in range(len(quantities))if i%2 != 0]


        # Iterate to check if any Item name or quantity is empty
        for selected_item, quantity in zip(selected_items,quantities):

            # Check if the item comboBox is empty
            if(selected_item.currentText()==""):
                selected_item.setFocus()
                QtWidgets.QToolTip.showText(selected_item.mapToGlobal(QtCore.QPoint(0,30)),"Select item")
                return

            # Check if the quantity field is empty
            elif(quantity.text()==""):
                quantity.setFocus()
                QtWidgets.QToolTip.showText(quantity.mapToGlobal(QtCore.QPoint(0,30)),"Enter quantity")
                return      

        # Iterate to add customer's orders_model
        items_quantities = tuple([(selected_item.currentText(), quantity.text()) for selected_item, quantity in zip(selected_items, quantities)])
        data = [(self.offer_name_txt.text(), self.offer_price_txt.text()),items_quantities]

        ret = self.offers_model.addOffer(data)
        
        if(ret):
            # Delete shift registring panel 
            self.clearLayout(self.offers_items_frame.layout())

            # Resest employee numbers variable
            self._item_number = 0

            # Insert an empty order
            self.plusOffer()

            # Reset offer name and price textboxes
            self.offer_name_txt.setText('')
            self.offer_price_txt.setText('')

            
            # Reinitial items
            self.initialSettingsComboBoxs()

            # Redisplay offers
            self.showOffers()

    def removeOffer(self):
        """Remove offer/s from offers table"""
        # Get selected rows and their indexs
        rows_indices = self.offers_tableView.selectionModel().selectedRows()
    
        if (len(rows_indices) <= 0):
            QtWidgets.QToolTip.showText(self.offer_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select offer/s")
        
        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                # Create a worker thread
                worker = Worker(self.daily_conn, 'Offers', 'offer_id', rows_indices)

                # Disable remove button

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.offers_model.submitAll())
                worker.finished.connect(lambda : self.offers_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Offers', column = 'offer_id', db1 = self.daily_conn))
                worker.finished.connect(self.initialSettingsComboBoxs)
                worker.finished.connect(self.Completers)
                
                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()

    def plusOffer(self):
        """Adding new order to orders list"""
        self._item_number += 1

        # Create register order panel 
        offer_frame = CustomAddOfferFrame(self._item_number, self.offers_items_frame)
        
        # Add the container frame into parent frame layout
        self.verticalLayout_45.addWidget(offer_frame)
        
        # pass on items comboBox
        offer_frame.offer_item_comboBox.addItems([''] + retrieveItemNames(db = self.daily_conn))

        # Connect each comboBox with its label.
        offer_frame.offer_item_comboBox.currentTextChanged['QString'].connect(lambda : self.isOfferItemExists(offer_frame.offer_item_comboBox, offer_frame.offer_quantity_txt))
       

        # pass on delete buttons to connect it with its frame.
        offer_frame.delete_btn.clicked.connect(lambda : self.deleteOfferFrame(offer_frame))
    
    def isOfferItemExists(self, selected_item, quntity_txtBox):
        # Check if the item is selected previously 
        selected_items = [emp.currentText() for emp in self.offers_items_frame.findChildren(QtWidgets.QComboBox)]

        if(retrieveItemId(selected_item.currentText(), db = self.daily_conn) is None):
            QtWidgets.QToolTip.showText(selected_item.mapToGlobal(QtCore.QPoint(0,30)),"Not found")
            selected_item.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 255, 0);\n"
            "}"
            )
            quntity_txtBox.setEnabled(False)

        # if the selected item is null
        elif(selected_item.currentText()==''):
            selected_item.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 170, 0);\n"
            "}"
            )
            quntity_txtBox.setEnabled(False)
            

        # Check if the selected item is previously existed
        elif(selected_items.count(selected_item.currentText()) > 1):
            QtWidgets.QToolTip.showText(selected_item.mapToGlobal(QtCore.QPoint(0,30)),"Selected previously")
            selected_item.setStyleSheet(
            "border-width: 0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 0, 0);")

            self.offer_add_btn.setEnabled(False)

            # Disable editing quantity text 
            quntity_txtBox.setEnabled(False)
            quntity_txtBox.setText('')
        
          
        # If the selected item is not selected previously
        else:
            selected_item.setStyleSheet(
            "border-width: 0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 170, 0);")
            
            self.offer_add_btn.setEnabled(True)
            quntity_txtBox.setEnabled(True)

    def deleteOfferFrame(self, frame):
        """Delete frame after clicking"""

        # Keep at least one order that shouldn't be deleted
        if len(self.offers_items_frame.findChildren(QtWidgets.QComboBox)) > 1:

            sip.delete(frame) 

            # decrement order number
            self._item_number -=1

            # Reset item labels name
            _translate = QtCore.QCoreApplication.translate
            labels = [label for label in self.offers_items_frame.findChildren(QtWidgets.QLabel) if(label.objectName() == 'item_number_lbl')]
            for num, label in zip(range(1, self._item_number+1),labels): 
                label.setText(_translate("MainWindow", f"Item : {num} "))
    
    def clearOffersFitlers(self):
        self.offers_sort_model.setItemNameFilter('')
        self.offers_sort_model.setDateFilter('')

    ###############
    # Employees #
    ###############
    # def showEmployees2(self):
    #     """Show all employees from Supervisors table"""
    #     emp_detail_frame = CustomEmployeeFrame(self.employees_frame)
    #     self.gridLayout_12.addWidget(1,1,1,1)

        
    #     self.setViewModel()

    #     # # If Daily panel is opend then we can make change to date (add, edit, ...)
    #     if(self._SETTINGS_TABLES_FLAG == True):
    #         self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'employees_properties_panel'))

    #     self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'tab'))
    #     self.toggleStackWidget(75, True)


    def showEmployees(self):
        """Show all employees from Supervisors table"""
        
        self.setViewModel()

        # # If Daily panel is opend then we can make change to date (add, edit, ...)
        if(self._SETTINGS_TABLES_FLAG == True):
            self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'employees_properties_panel'))

        self.employees_tableView.setModel(self.employees_sort_model)
        # self.employees_tableView.resizeColumnsToContents()

        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'employees_tab'))
        self.toggleStackWidget(75, True)

    def addEmployees(self):
        """Add new employee to Supervisors table"""
        # Check if the supervisor name field is empty
        if(self.employee_name_txt.text()==""):
            self.employee_name_txt.setFocus()
            QtWidgets.QToolTip.showText(self.employee_name_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter supervisor name")
            return
        
        # Check if the supervisor job type button is checked
        if(self.employee_manager_btn.isChecked()==False and self.employee_worker_btn.isChecked()==False):
            self.employee_worker_btn.setFocus()
            QtWidgets.QToolTip.showText(self.employee_worker_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select job type")
            return

        # Check if the supervisor username field is empty
        if(self.employee_username_txt.text()==""):
            self.employee_username_txt.setFocus()
            QtWidgets.QToolTip.showText(self.employee_username_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter username")
            return
        
        # Check if the supervisor password field is empty
        if(self.employee_password_txt.text()==""):
            self.employee_password_txt.setFocus()
            QtWidgets.QToolTip.showText(self.employee_password_txt.mapToGlobal(QtCore.QPoint(0,30)),"Enter password")
            return

        if(self.employee_manager_btn.isChecked()==True):
            job_type = self.employee_manager_btn.text()
        else:
            job_type = self.employee_worker_btn.text()
        
        
        data = [
            self.employee_name_txt.text().strip(),
            self.employee_gender_comboBox.currentText(),
            job_type,
            self.employee_username_txt.text().strip(), 
            self.employee_password_txt.text().strip()
        ]
    
        ret = self.employees_model.addEmployee(data)

        if(ret):
            # Reinitial items
            self.initialSettingsComboBoxs()

            # Resest text
            self.employee_name_txt.setText('')
            self.employee_username_txt.setText('')
            self.employee_password_txt.setText('')


        else:
            QtWidgets.QMessageBox.warning(
            self,
            "Info",
            "Record exists...!",
            QtWidgets.QMessageBox.Ok ,
            )   

    def removeEmployees(self):
        """Remove employee/s from Supervisors table"""
        # Get selected rows and their indexs
        rows_indices = self.employees_tableView.selectionModel().selectedRows()
    
        if (len(rows_indices) <= 0):
            QtWidgets.QToolTip.showText(self.employee_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select supervisor/s")
        
        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                # Create a worker thread
                worker = Worker(self.daily_conn, 'Supervisors', 'supervisor_id', rows_indices)

                # Disable remove button

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.employees_model.submitAll())
                worker.finished.connect(lambda : self.employees_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Supervisors', column = 'supervisor_id', db1 = self.daily_conn))
                worker.finished.connect(self.initialSettingsComboBoxs)
                worker.finished.connect(self.Completers)
                
                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()

    def clearEmployeesFilters(self):
        self.employees_employee_name_filter_txt.setText('')
        self.employees_job_type_filter_comberoBox.setCurrentText('')


    ##########
    # Shifts #
    ##########
    def showShifts(self):
        """Show all shifts from shifts table"""
        
        self.setViewModel()

        if(self._SETTINGS_TABLES_FLAG == True):
            self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'shifts_properties_panel'))

        self.shifts_tableView.setModel(self.shifts_sort_model)

        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'shifts_tab'))
        self.toggleStackWidget(75, True)

        self.showShiftsEmployees()

    def showShiftsEmployees(self):
        """Display available shifts and related employees"""

        self.shifts_supervisors_model = HierarcicalShiftsSupervisorsModel(self.daily_conn)
        self.shifts_supervisors_treeView.setHeaderHidden(True)
        self.shifts_supervisors_treeView.setModel(self.shifts_supervisors_model) 
        self.shifts_supervisors_treeView.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        
        # selection_model = self.shifts_supervisors_treeView.selectionModel()
        # # selection_model.selectionChanged.connect(self.selectionChangedSlot)
            
    def addShift(self):
        """Add new shift with related employees"""
        
        # Take children of order register panel 
        # Item and quantity
        combo_selected_employees = self.shift_employees_frame.findChildren(QtWidgets.QComboBox)

        # Iterate to check if any Item name or quantity is empty
        for combo_selected_employee in combo_selected_employees:
            # Check if the item comboBox is empty
            if(combo_selected_employee.currentText()==""):
                combo_selected_employee.setFocus()
                QtWidgets.QToolTip.showText(combo_selected_employee.mapToGlobal(QtCore.QPoint(0,30)),"Select employee")
                return   

        # Iterate to add customer's orders_model
        data = [combo_selected_employee.currentText() for combo_selected_employee in combo_selected_employees]

        self.shifts_model.addShift(data)
        
        # Delete shift registring panel 
        self.clearLayout(self.shift_employees_frame.layout())

        # Resest employee numbers variable
        self._employee_number = 0

        # Insert an empty order
        # self.plusEmployee()
        
        # Reinitial items
        self.initialSettingsComboBoxs()

        self.showShiftsEmployees()

    def plusEmployee(self):
        """Adding new order to orders list"""
        self._employee_number += 1

        # Create register order panel 

        employee_frame = QtWidgets.QFrame(self.shift_employees_frame)
        employee_frame.setLayoutDirection(QtCore.Qt.LeftToRight)
        employee_frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        employee_frame.setFrameShadow(QtWidgets.QFrame.Raised)
        employee_frame.setObjectName("employee_frame")

        # Add layout to the panel
        horizontalLayout = QtWidgets.QHBoxLayout(employee_frame)
        horizontalLayout.setContentsMargins(0, 0, 0, 0)
        horizontalLayout.setSpacing(11)
        horizontalLayout.setObjectName("horizontalLayout")

        # Add label to the panel for item number
        employee_name_lbl = QtWidgets.QLabel(employee_frame)
        employee_name_lbl.setStyleSheet("QLabel{\n"
            "\n"
            "color: rgb(255, 255, 255);\n"
            "}")
        employee_name_lbl.setObjectName("employee_name_lbl")

        _translate = QtCore.QCoreApplication.translate
        employee_name_lbl.setText(_translate("MainWindow", f"Employee : {self._employee_number} "))
        horizontalLayout.addWidget(employee_name_lbl)

        # Add item comboBox to the layout 
        employee_comboBox = QtWidgets.QComboBox(employee_frame)
        employee_comboBox.setMinimumSize(QtCore.QSize(90, 40))
        employee_comboBox.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 170, 0);\n"
            "}"
            )
        employee_comboBox.setEditable(True)
        employee_comboBox.setObjectName("employee_comboBox")
        employee_comboBox.setEditable(True)
        horizontalLayout.addWidget(employee_comboBox) 

        # Add delete current input order button to the layout
        delete_btn = QtWidgets.QPushButton(employee_frame)
        delete_btn.setMinimumSize(QtCore.QSize(50, 50))
        delete_btn.setMaximumSize(QtCore.QSize(50, 50))
        delete_btn.setStyleSheet(
            "QPushButton{\n"
            "\n"
            "    border-style: solid;\n"
            "    border-width: 4px 4px 4px 4px;\n"
            "    border-radius: 25px;    \n"
            "    border-color: rgb(174, 174, 174);\n"
            "    background-color: rgb(255, 255, 255);\n"
            "    image: url(:/icons/icons/letter-x2.svg);\n"
            "    padding:9px;\n"
            "}\n"
            "QPushButton:hover{\n"
            "    image: url(:/icons/icons/letter-x1.svg);\n"

           
            "}\n"
            "QPushButton:pressed{\n"
            "    padding:14px;\n"
            "}\n"
            "")

        # icon = QtGui.QIcon()
        # icon.addPixmap(QtGui.QPixmap(":/icons/icons/cancel.svg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        # delete_btn.setIcon(icon)
        # delete_btn.setIconSize(QtCore.QSize(20, 20))
        horizontalLayout.addWidget(delete_btn)
        
        # Add the container frame into parent frame layout
        self.verticalLayout_98.addWidget(employee_frame)
        
        # pass on items comboBox
        employee_comboBox.addItems([''] + retrieveEmployeesNames(db = self.daily_conn))
        
        # Connect each comboBox with its label.
        employee_comboBox.currentTextChanged['QString'].connect(lambda : self.isEmpExists(employee_comboBox))
       

        # pass on delete buttons to connect it with its frame.
        delete_btn.clicked.connect(lambda : self.deleteEmployeeFrame(employee_frame))
            
    def isEmpExists(self, employee_comboBox):
        # Check if the item is selected previously 
        combo_selected_employees = [emp.currentText() for emp in self.shift_employees_frame.findChildren(QtWidgets.QComboBox)]
        if(employee_comboBox.currentText()==''):
            employee_comboBox.setStyleSheet(
            "QComboBox{\n"
            "border-width:0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 170, 0);\n"
            "}"
            )

    
        # Check if the employee selected previously
        elif(combo_selected_employees.count(employee_comboBox.currentText()) > 1):
            QtWidgets.QToolTip.showText(employee_comboBox.mapToGlobal(QtCore.QPoint(0,30)),"Selected previously")
            employee_comboBox.setStyleSheet(
            "border-width: 0px 0px 6px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 0, 0);")
            self.shift_add_btn.setEnabled(False)
          
        # If the employee not selected previously we
        else:
            employee_comboBox.setStyleSheet(
            "border-width: 0px 0px 4px 0px;\n"
            "border-style: solid;\n"
            "border-radius:0px;\n"
            "border-color: rgb(255, 170, 0);")
            
            self.shift_add_btn.setEnabled(True)
  
    def deleteEmployeeFrame(self, frame):
        """Delete frame after clicking"""

        # Keep at least one order that shouldn't be deleted
        if len(self.shift_employees_frame.findChildren(QtWidgets.QComboBox)) > 1:

            sip.delete(frame) 

            # decrement order number
            self._employee_number -=1

            # Reset item labels name
            _translate = QtCore.QCoreApplication.translate
            labels = [label for label in self.shift_employees_frame.findChildren(QtWidgets.QLabel) if(label.objectName() == 'employee_name_lbl')]
            for num, label in zip(range(1, self._employee_number+1),labels): 
                label.setText(_translate("MainWindow", f"Employee : {num} "))
    
    def removeShifts(self):
        """Remove shifts from Shifts table"""
        # Get selected rows and their indexs
        rows_indices = self.shifts_tableView.selectionModel().selectedRows()
    
        if (len(rows_indices) <= 0):
            QtWidgets.QToolTip.showText(self.employee_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select shift/s")
        
        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                # Create a worker thread
                worker = Worker(self.daily_conn, 'Shifts', 'shift_id', rows_indices)

                # Disable remove button

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.shifts_model.submitAll())
                worker.finished.connect(lambda : self.shifts_model.select())
                worker.finished.connect(lambda : resetCounting(table = 'Shifts', column = 'shift_id', db1 = self.daily_conn))
                worker.finished.connect(self.initialSettingsComboBoxs)
                worker.finished.connect(self.showShiftsEmployees)
                worker.finished.connect(self.reinitialAvailableEmpComboBox)
                worker.finished.connect(self.Completers)
                
                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()
    
    def reinitialAvailableEmpComboBox(self):
        
        employees_comboBox = self.shift_employees_frame.findChildren(QtWidgets.QComboBox)
        for employee_comboBox in employees_comboBox:
            employee_comboBox.addItems([''] + retrieveEmployeesNames(db = self.daily_conn))

    def startShift(self):
        """Start selected shift and stop other shifts"""

        # Get selected rows and their indexs
        index = self.shifts_tableView.selectionModel().currentIndex()
        
        shift_id = index.sibling(index.row(), self.shifts_model.fieldIndex('shift_id')).data()
        shift_state = index.sibling(index.row(), self.shifts_model.fieldIndex('shift_state')).data()

        if (shift_id == None):
            QtWidgets.QToolTip.showText(self.shift_start_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select shift/s")
        
        elif(shift_state != 'Active'):
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Starting Alert",
                "Start selected shift and stop others, Confirm?",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                startShift(shift_id, db=self.daily_conn)
                self.isAnyShiftActive()
                self.showShifts()

    def finishShift(self):
        """Stop selected shift"""

        # Get selected rows and their indexs
        index = self.shifts_tableView.selectionModel().currentIndex()
        id = index.sibling(index.row(), self.shifts_model.fieldIndex('shift_id')).data()
    
        if (id == None):
            QtWidgets.QToolTip.showText(self.shift_start_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select shift/s")
        
        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Stop Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                finishShift(id, db=self.daily_conn)
                self.showShifts()

    def clearShiftsFitlers(self):
        self.shifts_sort_model.setDateFilter('')


    ###########
    # Reports #
    ###########
    def showReports(self):
        """Show Reports tables"""


        # If Daily panel is opend then we can make change to date (add, edit, ...)
        if(self._SETTINGS_TABLES_FLAG == True):
            self.stackedWidget.setCurrentWidget(self.stackedWidget.findChild(QtWidgets.QWidget, 'reports_properties_panel'))
            self.toggleStackWidget(75, True)
            self.showYMDs(self.daily_conn, 'Reports', 'date')

        
        self.reports_tableView.setModel(self.reports_sort_model)

        self.report_search_type_comboBox.setCurrentText('')

        self.tabWidget.setCurrentWidget(self.tabWidget.findChild(QtWidgets.QWidget, 'reports_tab'))

    def reportType(self, report_type):
        """Determine report type"""
        self.reports_model.showReports(filter = report_type)

    def removeReport(self):
        rows_indices = self.reports_tableView.selectionModel().selectedRows()
        
        if (len(rows_indices) <= 0):
            QtWidgets.QToolTip.showText(self.order_remove_btn.mapToGlobal(QtCore.QPoint(0,30)),"Select report/s")

        else:
            messageBox = QtWidgets.QMessageBox.warning(
                self,
                "Deleting Alert",
                "Confirm",
                QtWidgets.QMessageBox.Ok | QtWidgets.QMessageBox.Cancel,
                QtWidgets.QMessageBox.Cancel
            )
            if messageBox == QtWidgets.QMessageBox.Ok:
                

                worker = ReportsWorker(self.daily_conn, 'Reports', 'date', rows_indices)

                # Disable remove button
                self.report_remove_btn.setEnabled(False)

                worker.finished.connect(worker.deleteLater)
                worker.finished.connect(lambda : self.report_remove_btn.setEnabled(True))
                worker.finished.connect(lambda : self.reportType(self.report_search_type_comboBox.currentText()))
                worker.finished.connect(lambda : updateReports(db= self.daily_conn))
                
                
                # Start the thread
                worker.start()

                # Create progressBar
                prog = PorgressBarDialog()
                prog.progressBar.setMaximum(len(rows_indices))
                prog.progressBar.setMinimum(0)
                worker.row_num_changed.connect(prog.progressBar.setValue)
                worker.finished.connect(prog.close)
                prog.exec()

    def clearReportsFitlers(self):
        self.reports_sort_model.setDateFilter('')

    # def drawGraphReport(self):
        
        # if(self.report_search_type_comboBox.currentText() !=''):
        #     date, total = [], []
        #     for row in range(self.reports_model.rowCount()):
        #         date.append(self.reports_model.data(self.reports_model.index(row,0), Qt.DisplayRole))
        #         total.append(self.reports_model.data(self.reports_model.index(row,self.reports_model.columnCount()-1), Qt.DisplayRole))

        #     graph = GraphWidget(self)
        #     graph.draw(date, total)
        #     graph.show()
    
    ######################
    # Subscriptions fees #
    ######################
    def openPriceDialog(self):
        """Open change price dialog"""

        self.slideErrorFrame(0, '', True)
        
        price_diag = PriceDialog(db = self.daily_conn)
        if(price_diag.exec() == QtWidgets.QDialog.Accepted):
            """Change subscription price"""
            subs_type = price_diag.data[0]
            price = price_diag.data[1]
            changeSubsCost(price, subs_type, self.daily_conn)    

    def isAnyFeeZero(self):
        ret = checkFeesValue(db = self.daily_conn)
        if(ret == 0):
            self.slideErrorFrame(0, '',True)
            return True

        elif(ret > 0 ):
            self.slideErrorFrame(100, 'Please set all fees value before starting', True)
            return False

    ############
    # Archieve #
    ############
    def showYMDs(self, db, table, field):
        """Display available years, months, and days from Archive"""

        self.date_model = HierarcicalDateModel(db, table, field)
        self.date_treeView.setModel(self.date_model) 
        self.date_treeView.setHeaderHidden(True)
        self.date_treeView.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        
        self.date_treeView.selectionModel().selectionChanged.connect(self.selectionChangedSlot)
            
    def selectionChangedSlot(self, newSelection : QtCore.QItemSelection, oldSelection : QtCore.QItemSelection):
        """pyqtSlot to get the selection text"""
        # get the text of the selected item
        index = self.date_treeView.selectionModel().currentIndex()
        selectedText = index.data(Qt.DisplayRole)

        # find out the hierarchy level of the selected item
        # hierarchyLevel = 1
        # seekRoot = index
        # while (seekRoot.parent() !== QtCore.QModelIndex()):
        #     seekRoot = seekRoot.parent()
        #     hierarchyLevel += 1
        
        # showString = "{0}, Level {1}".format(selectedText, hierarchyLevel)

        date = [index.parent().parent().data(Qt.DisplayRole), index.parent().data(Qt.DisplayRole), selectedText]
        date = '-'.join([d for d in date if d != None])

        self.filterByDate(date)

    @QtCore.pyqtSlot(str)
    def filterByDate(self, date):
        """Get specific customer subscription type from daily table"""

        # # Formate date to English form 
        # locale  = QtCore.QLocale(QtCore.QLocale.English, QtCore.QLocale.AmericanSamoa)

        # # Convert fromated date to string
        # english_date = locale.toString(date, "yyyy-MM-dd")

        if(self._ARCHIVE_TABLES_FLAG):
            self.daily_customers_sort_model.setDateFilter(date)
            self.monthly_customers_sort_model.setDateFilter(date)
            self.orders_sort_model.setDateFilter(date)

        elif(self._SETTINGS_TABLES_FLAG):
            if(self.report_search_type_comboBox.currentText() in ['شهري','معدل الزبائن']):
                self.reports_sort_model.setDateFilter(date.split('-')[-1])
            else:
                self.reports_sort_model.setDateFilter(date)


    #################################
    # Exports tables to Excel files #
    #################################
    def exportTable(self, model):
        """Export table to excel file"""
        table_name = ''
        if(model != self.reports_model):
            table_name = model.tableName()
        else:
            table_name = 'Report'
        
        # get current date
        currnet_date = datetime.datetime.now().strftime("%Y-%m-%d") # or %B %d, %Y
        
        if(self._DAILY_TABLES_FLAG):
            type = 'Daily'
        elif(self._ARCHIVE_TABLES_FLAG):
            type = 'Archive'
        elif(self._SETTINGS_TABLES_FLAG):
            type = 'Daily'

        
        # get the file name which is saved
        fname, _ = QtWidgets.QFileDialog.getSaveFileName(self,caption='Save file',
                    directory = os.path.join(os.getcwd(),f"{type}-{table_name} ({currnet_date}).xlsx"),
                    filter="XLSX files (*.xlsx)")
                
        # check if the file extension is an Excel file
        reg = re.compile(r"\.xlsx$")

        if (reg.search(fname) is not None):

            # check if the work book exists

            # create excel file
            excel_file = openpyxl.Workbook(write_only=False)
            
            # # create sheet inside excel file
            # sheet = excel_file.active
            
            # create sheet inside excel file
            sheet = excel_file.create_sheet(title = f'{currnet_date}')

            # create columns header
            for col in range(model.columnCount()):
                column_header = model.headerData(col,Qt.Horizontal,Qt.EditRole)
                sheet.cell(1, col+1).value = column_header
            

            # add data to cells
            for row in range(model.rowCount()):
                for col in range(model.columnCount()):
                    value = model.data(model.index(row,col), Qt.DisplayRole)
                    sheet.cell(row+2, col+1).value = value

            excel_file.save(fname)
            excel_file.close()
